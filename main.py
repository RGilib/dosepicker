#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Nov 21 08:31:34 2025

@author: giliberr@menarini.net
"""

# main.py
import os, re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from typing import Optional, List, Tuple, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

ALLOWED_ANY_DATA = [("Data files", "*.csv *.tsv *.xlsx *.xls *.json"), ("All files", "*.*")]

# ----------------- header/IO helpers -----------------
def normalize_header(name: str) -> str:
    s = str(name).replace("\u00A0", " ")
    s = re.sub(r"[\s\u00A0]+", " ", s.strip())
    return s.lower()

def nospace_key(s: str) -> str:
    return s.replace(" ", "")

VISIT_CYCLE_DAY_RE = re.compile(r'(?i)\bC\s*(\d+)\s*D\s*(\d+)\b')

def parse_cycle_day(visit_str: str):
    """Return (cycle:int, day:int) if visit has C#D# pattern, else None."""
    if visit_str is None or pd.isna(visit_str):
        return None
    s = str(visit_str).replace("\u00A0", " ").strip()
    m = VISIT_CYCLE_DAY_RE.search(s)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))

def find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    lut = {nospace_key(normalize_header(c)): c for c in df.columns}
    for cand in candidates:
        k = nospace_key(normalize_header(cand))
        if k in lut:
            return lut[k]
    return None

def read_table(path: str) -> pd.DataFrame:
    p = path.lower()
    try:
        if p.endswith(".csv"):
            return pd.read_csv(path, dtype=str)
        if p.endswith((".tsv", ".tab")):
            return pd.read_csv(path, sep="\t", dtype=str)
        if p.endswith((".xlsx", ".xls")):
            return pd.read_excel(path, dtype=str)
        if p.endswith(".json"):
            df = pd.read_json(path, orient="records")
            return df.applymap(lambda v: None if pd.isna(v) else str(v))
    except Exception:
        pass
    try:
        return pd.read_csv(path, dtype=str)
    except Exception:
        return pd.read_excel(path, dtype=str)

def _clean_cell(x):
    if pd.isna(x): return pd.NA
    s = str(x).replace("\u00A0", " ").strip()
    return pd.NA if s == "" else s

def _keyize_id(x) -> str:
    if pd.isna(x): return ""
    s = str(x).replace("\u00A0", " ").strip()
    if s.isdigit():
        s = s.lstrip("0") or "0"
    return s

def _to_percent(value):
    if pd.isna(value): return pd.NA
    s = str(value).replace("\u00A0", " ").strip()
    if s == "": return pd.NA
    s = s.replace("%", "").replace(",", ".")
    try:
        v = float(s)
    except Exception:
        return pd.NA
    if v <= 1.0:
        v *= 100.0
    if v < 0: v = 0.0
    if v > 1000: v = 100.0
    return v

# ----------------- visit logic & cohorting -----------------
EXCLUDED_VISITS = {"unscheduled visit 1", "unscheduled visit 2", "."}
BASELINE_NAMES  = {"baseline", "screen", "screening"}

def _norm_visit(v) -> str:
    if pd.isna(v): return ""
    return str(v).replace("\u00A0", " ").strip().lower()

def is_excluded_visit(v) -> bool:
    return _norm_visit(v) in EXCLUDED_VISITS

def is_baseline(v) -> bool:
    return _norm_visit(v) in BASELINE_NAMES

def has_valid_timepoint_nonbaseline(visit_series: pd.Series) -> bool:
    """At least one non-baseline visit that matches C#D# and is not excluded."""
    if visit_series is None or visit_series.empty:
        return False
    for v in visit_series.dropna():
        if is_excluded_visit(v): 
            continue
        if is_baseline(v):
            continue
        if parse_cycle_day(v) is not None:
            return True
    return False

def eligible_subjects(df: pd.DataFrame, subj_col: str, visit_col: str) -> set:
    tmp = df[[subj_col, visit_col]].copy()
    tmp["_KEY"] = tmp[subj_col].map(_keyize_id)
    tmp = tmp[tmp["_KEY"] != ""]
    tmp = tmp[~tmp[visit_col].map(is_excluded_visit)]
    g = tmp.groupby("_KEY")[visit_col]
    has_base   = g.apply(lambda s: any(is_baseline(x) for x in s)).astype(bool)
    has_follow = g.apply(lambda s: any((not is_baseline(x)) for x in s)).astype(bool)
    ok = has_base & has_follow
    return set(ok[ok].index.tolist())

def cohorts_from_dose(dose_df: pd.DataFrame, col_patient: str, col_group: str, valid_keys: set) -> Dict[str, List[str]]:
    keys   = dose_df[col_patient].map(_keyize_id)
    groups = dose_df[col_group].map(_clean_cell)
    tmp = pd.DataFrame({"_KEY": keys, "__group__": groups})
    tmp = tmp.dropna(subset=["__group__"])
    tmp = tmp[tmp["_KEY"].isin(valid_keys)]
    out: Dict[str, List[str]] = {}
    for grp, sub in tmp.groupby("__group__"):
        ordered, seen = [], set()
        for k in sub["_KEY"]:
            if k not in seen:
                seen.add(k); ordered.append(k)
        if ordered:
            out[str(grp)] = ordered
    return out

def _cohort_sort_key(name: str):
    m = re.search(r'(\d+)', str(name))
    return (int(m.group(1)) if m else float('inf'), str(name).lower())

def format_cohort_header(name: str) -> str:
    m = re.fullmatch(r"\s*tag\s*([0-9]+)\s*", name, flags=re.I)
    if m:
        return f"TAG {m.group(1)} mgc"
    return name

def iqr(values: List[float]) -> Optional[Tuple[float, float]]:
    if not values: return None
    vals = sorted(values)
    n = len(vals)
    def _q(p):
        idx = (p*(n+1))/4
        i = int(idx)
        f = idx - i
        if i <= 0: return vals[0]
        if i >= n: return vals[-1]
        return vals[i-1] + f*(vals[i]-vals[i-1])
    return (round(_q(1),1), round(_q(3),1))

# ----------------- visit ordering -----------------
def build_ordered_visits(input_df: pd.DataFrame,
                         col_subject: str,
                         col_visit: str,
                         col_bm: str,
                         col_pb: str,
                         col_cd123: str,
                         ok_keys: set) -> list[str]:
    """
    Screening first (if any data). Then only visits with C#D# pattern,
    ordered by (cycle, day). Drop timepoints with zero usable data
    (no BM blasts and no BM*CD123 product across eligible subjects).
    """
    def _visit_has_data(visit_label: str) -> bool:
        sub = input_df[~input_df[col_visit].map(is_excluded_visit)].copy()
        sub = sub[sub[col_subject].map(_keyize_id).isin(ok_keys)]
        sub = sub[sub[col_visit].astype(str).str.strip().str.lower() == _norm_visit(visit_label)]
        has_bm = pd.to_numeric(sub[col_bm].map(_to_percent), errors="coerce").notna().any()
        bm = pd.to_numeric(sub[col_bm].map(_to_percent), errors="coerce")
        cd = pd.to_numeric(sub[col_cd123].map(_to_percent), errors="coerce")
        prod = (bm * (cd/100.0))
        has_prod = prod.notna().any()
        return bool(has_bm or has_prod)

    allowed = input_df[~input_df[col_visit].map(is_excluded_visit)][[col_visit]].dropna()
    unique_visits = {str(v).strip() for v in allowed[col_visit].tolist()}

    visit_list = []
    if any(_norm_visit(v) in BASELINE_NAMES for v in unique_visits):
        if _visit_has_data("Screening"):
            visit_list.append("Screening")

    coded = []
    for v in unique_visits:
        parsed = parse_cycle_day(v)
        if parsed is None:
            continue
        if not _visit_has_data(v):
            continue
        cycle, day = parsed
        coded.append((cycle, day, v))

    coded.sort(key=lambda t: (t[0], t[1], t[2].lower()))
    visit_list.extend([v for _, _, v in coded])
    return visit_list

# ----------------- mutation classification helpers -----------------
def _detect_mut_and_vaf_cols(df: pd.DataFrame) -> Tuple[List[str], List[str], List[int]]:
    """
    Detect TP53DNA{n} and TP53VAF{n} actual column names and the sorted indices n.
    """
    lut = {nospace_key(normalize_header(c)): c for c in df.columns}
    idxs = set()
    for norm in lut:
        m = re.fullmatch(r"tp53dna(\d+)", norm or "")
        if m: idxs.add(int(m.group(1)))
    vaf_idxs = set()
    for norm in lut:
        m = re.fullmatch(r"tp53vaf(\d+)", norm or "")
        if m: vaf_idxs.add(int(m.group(1)))
    idxs = sorted(list(idxs & vaf_idxs))
    dna_cols = [lut[f"tp53dna{i}"] for i in idxs]
    vaf_cols = [lut[f"tp53vaf{i}"] for i in idxs]
    return dna_cols, vaf_cols, idxs

def _cohort_mutations_df(input_df: pd.DataFrame,
                         keys: set,
                         col_subject: str,
                         col_visit: str,
                         dict_path: Optional[str]) -> pd.DataFrame:
    """
    Build mutation table for given cohort keys, using your gating:
      - subjects with ≥1 valid non-baseline C#D# visit
      - include all rows for those subjects
      - extract all (TP53DNAn, TP53VAFn) pairs, VAF as percent
      - left-merge with TP53 dictionary (index_col=0) if provided
    Output columns: ['mutation_id','vaf','Effect','Hotspot','_SUBJECT_KEY']
    """
    if not keys:
        return pd.DataFrame(columns=["mutation_id","vaf","Effect","Hotspot","_SUBJECT_KEY"])

    sub = input_df[input_df[col_subject].map(_keyize_id).isin(keys)].copy()
    sub = sub[~sub[col_visit].map(is_excluded_visit)]
    keep_keys = set()
    for k, g in sub.groupby(sub[col_subject].map(_keyize_id)):
        if has_valid_timepoint_nonbaseline(g[col_visit]):
            keep_keys.add(k)
    if not keep_keys:
        return pd.DataFrame(columns=["mutation_id","vaf","Effect","Hotspot","_SUBJECT_KEY"])

    to_check = input_df[input_df[col_subject].map(_keyize_id).isin(keep_keys)].copy()

    dna_cols, vaf_cols, idxs = _detect_mut_and_vaf_cols(input_df)
    records = []
    for idx, (dna_c, vaf_c) in zip(idxs, zip(dna_cols, vaf_cols)):
        pairs = to_check[[col_subject, dna_c, vaf_c]].dropna(subset=[dna_c, vaf_c])
        if pairs.empty:
            continue
        for _, row in pairs.iterrows():
            mut_id = str(row[dna_c]).strip()
            vaf = _to_percent(row[vaf_c])
            if pd.isna(vaf) or mut_id == "":
                continue
            records.append({
                "_SUBJECT_KEY": _keyize_id(row[col_subject]),
                "mutation_id": mut_id,
                "vaf": float(vaf)
            })

    mut_df = pd.DataFrame(records, columns=["_SUBJECT_KEY","mutation_id","vaf"])
    if mut_df.empty:
        return pd.DataFrame(columns=["mutation_id","vaf","Effect","Hotspot","_SUBJECT_KEY"])

    dict_df = None
    if dict_path and os.path.isfile(dict_path):
        try:
            dict_df = pd.read_csv(dict_path, dtype=str, index_col=0)
        except Exception:
            try:
                dict_df = pd.read_excel(dict_path, dtype=str, index_col=0)
            except Exception:
                dict_df = None

    if dict_df is not None and not dict_df.empty:
        dict_df = dict_df.copy()
        dict_df.index = dict_df.index.map(lambda x: "" if pd.isna(x) else str(x).strip())
        dict_df.columns = [str(c) for c in dict_df.columns]
        merged = mut_df.merge(dict_df, left_on="mutation_id", right_index=True, how="left")
    else:
        merged = mut_df.copy()
        if "Effect" not in merged.columns:
            merged["Effect"] = pd.NA
        if "Hotspot" not in merged.columns:
            merged["Hotspot"] = pd.NA

    merged["Effect"]  = merged["Effect"].astype(str).str.replace("\u00A0"," ").str.strip().str.lower()
    merged["Hotspot"] = merged["Hotspot"].astype(str).str.replace("\u00A0"," ").str.strip().str.lower()
    return merged

def _count_mean_text(df: pd.DataFrame, mask: pd.Series) -> str:
    if df is None or df.empty:
        return "0 (-)"
    sub = df[mask]
    if sub.empty:
        return "0 (-)"
    mean_v = round(float(sub["vaf"].mean()), 1)
    return f"{len(sub)} ({mean_v}%)"

# ----------------- core generator -----------------
def generate_fda_combo_excel(input_path: str, dose_path: str, tp53_dict_path: Optional[str], out_path: str):
    input_df = read_table(input_path)
    dose_df  = read_table(dose_path)

    col_subject = find_col(input_df, ["SUBJECT"]) or "SUBJECT"
    col_visit   = find_col(input_df, ["VISIT"])   or "VISIT"
    col_bm      = find_col(input_df, ["BMBLASTS","BM BLASTS"])
    col_pb      = find_col(input_df, ["PBBLASTS","PB BLASTS"])  # not used for product
    col_cd123   = find_col(input_df, ["CD123"])
    if not col_bm or not col_cd123:
        raise ValueError("Missing required columns in INPUT: need BMBLASTS (or BM BLASTS) and CD123.")
    if not col_pb:
        col_pb = col_bm  # harmless fallback

    col_patient = find_col(dose_df, ["Patient Number"])
    col_group   = find_col(dose_df, ["Treatment Group"])
    if not col_patient or not col_group:
        raise ValueError("Missing required columns in DOSE DICT: need 'Patient Number' and 'Treatment Group'.")

    # eligible subjects (baseline + ≥1 post-baseline; excluding Unscheduled/.)
    ok_keys = eligible_subjects(input_df, col_subject, col_visit)
    cohorts = cohorts_from_dose(dose_df, col_patient, col_group, valid_keys=ok_keys)
    cohort_names = sorted(list(cohorts.keys()), key=_cohort_sort_key)

    # ordered visit list to display
    visit_list = build_ordered_visits(
        input_df=input_df,
        col_subject=col_subject,
        col_visit=col_visit,
        col_bm=col_bm,
        col_pb=col_pb,
        col_cd123=col_cd123,
        ok_keys=ok_keys,
    )
    n_vis = len(visit_list)

    def values_for(df, keys: set, visit_name: str, value_col: str) -> List[float]:
        sub = df[df[col_subject].map(_keyize_id).isin(keys)]
        sub = sub[~sub[col_visit].map(is_excluded_visit)]
        sub = sub[sub[col_visit].astype(str).str.strip().str.lower() == _norm_visit(visit_name)]
        vals = pd.to_numeric(sub[value_col].map(_to_percent), errors="coerce")
        return [float(x) for x in vals.dropna().tolist()]

    def mean_safe(vals: List[float]) -> Optional[float]:
        if not vals: return None
        return round(sum(vals)/len(vals), 1)

    def fold_vs_baseline_mean(keys: set, visit_name: str) -> str:
        sub_all = input_df[~input_df[col_visit].map(is_excluded_visit)].copy()
        sub_all = sub_all[sub_all[col_subject].map(_keyize_id).isin(keys)]
        base_vals = sub_all.loc[sub_all[col_visit].map(is_baseline), col_bm]
        base_vals = pd.to_numeric(base_vals.map(_to_percent), errors="coerce").dropna()
        v_sub = sub_all[sub_all[col_visit].astype(str).str.strip().str.lower() == _norm_visit(visit_name)]
        v_vals = pd.to_numeric(v_sub[col_bm].map(_to_percent), errors="coerce").dropna()
        if base_vals.empty or v_vals.empty: return "-"
        base_mean = base_vals.mean()
        v_mean    = v_vals.mean()
        if base_mean <= 0 or v_mean <= 0: return "-"
        f = base_mean / v_mean
        return f"{round(f,1)}-fold"

    def max_fold_reduction_bm(keys: set) -> str:
        df = input_df[~input_df[col_visit].map(is_excluded_visit)].copy()
        df["_KEY"] = df[col_subject].map(_keyize_id)
        df = df[df["_KEY"].isin(keys)]
        df["is_base"] = df[col_visit].map(is_baseline)
        df["BM%"] = pd.to_numeric(df[col_bm].map(_to_percent), errors="coerce")
        best = 0.0
        for _, g in df.groupby("_KEY"):
            base = g.loc[g["is_base"], "BM%"].dropna()
            foll = g.loc[~g["is_base"], "BM%"].dropna()
            if base.empty or foll.empty: continue
            base_mean = base.mean()
            min_f = foll.min()
            if base_mean > 0 and min_f > 0:
                best = max(best, base_mean / min_f)
        return "-" if best <= 0 else f"{round(best,1)}-fold reduction"

    def mean_min_max_product(keys: set, visit_name: str) -> Tuple[Optional[float], Optional[float], Optional[float]]:
        # PRODUCT = BM * CD123
        sub = input_df[~input_df[col_visit].map(is_excluded_visit)].copy()
        sub = sub[sub[col_subject].map(_keyize_id).isin(keys)]
        sub = sub[sub[col_visit].astype(str).str.strip().str.lower() == _norm_visit(visit_name)]
        bm = pd.to_numeric(sub[col_bm].map(_to_percent), errors="coerce")
        cd = pd.to_numeric(sub[col_cd123].map(_to_percent), errors="coerce")
        prod = (bm * (cd/100.0)).dropna()
        if prod.empty:
            return (None, None, None)
        return (round(prod.mean(),1), round(prod.min(),1), round(prod.max(),1))

    def max_fold_reduction_product(keys: set) -> str:
        df = input_df[~input_df[col_visit].map(is_excluded_visit)].copy()
        df["_KEY"] = df[col_subject].map(_keyize_id)
        df = df[df["_KEY"].isin(keys)]
        df["is_base"] = df[col_visit].map(is_baseline)
        bm = pd.to_numeric(df[col_bm].map(_to_percent), errors="coerce")
        cd = pd.to_numeric(df[col_cd123].map(_to_percent), errors="coerce")
        df["PROD"] = (bm * (cd/100.0))
        best = 0.0
        for _, g in df.groupby("_KEY"):
            base_vals = g.loc[g["is_base"], "PROD"].dropna()
            follow    = g.loc[~g["is_base"], "PROD"].dropna()
            if base_vals.empty or follow.empty: continue
            base_mean = base_vals.mean()
            min_f = follow.min()
            if base_mean > 0 and min_f > 0:
                best = max(best, base_mean / min_f)
        return "-" if best <= 0 else f"{round(best,1)}-fold reduction"

    # ----------------- build workbook -----------------
    wb = Workbook()
    ws = wb.active
    ws.title = "FDA combo table_Blasts_TP53"

    def set_col_width(c, w):
        ws.column_dimensions[get_column_letter(c)].width = w

    # A1:D1 merged title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws["A1"].value = "Evaluable immunogenicity population includes all patients with a baseline assessment and at least 1 post-baseline assessment"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")

    # B3:C4 merged PD Assessment
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=3)
    ws["B3"].value = "PD Assessment"
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Cohort headers + Overall (sorted by TAG number)
    cols = []
    start_col = 4
    for i, grp in enumerate(cohort_names):
        col_idx = start_col + i
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        header = f"{format_cohort_header(grp)}\nN={len(cohorts[grp])}"
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cols.append(col_idx)

    overall_col = start_col + len(cohort_names)
    ws.merge_cells(start_row=3, start_column=overall_col, end_row=4, end_column=overall_col)
    ws.cell(row=3, column=overall_col, value=f"Overall\nN={len(ok_keys)}").alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cols.append(overall_col)

    # Section 1: BM blasts mean & IQR / fold vs baseline
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3)
    ws["B5"].value = "Frequency of BMB  % and range (IQR)"
    ws["B5"].alignment = Alignment(horizontal="center")

    for i, v in enumerate(visit_list, start=6):
        ws.cell(row=i, column=3, value=str(v))
        for col_idx, grp in zip(cols, cohort_names + ["__OVERALL__"]):
            keys = ok_keys if grp == "__OVERALL__" else set(cohorts[grp])
            vals = values_for(input_df, keys, v, col_bm)
            n_pat = len(vals)
            if is_baseline(v):
                if n_pat == 0:
                    text = "n=0"
                else:
                    m = mean_safe(vals)
                    q = iqr(vals)
                    text = f"n={n_pat}\n{m}%"
                    if q: text += f"\n({q[0]}–{q[1]})"
            else:
                if n_pat == 0:
                    text = "n=0"
                else:
                    m = mean_safe(vals)
                    fc = fold_vs_baseline_mean(keys, v)
                    text = f"n={n_pat}\n{m}%\n{fc}"
            ws.cell(row=i, column=col_idx, value=text).alignment = Alignment(wrap_text=True, vertical="top")

    # Max change in BM
    row_max_bm_title = 6 + n_vis
    ws.merge_cells(start_row=row_max_bm_title, start_column=2, end_row=row_max_bm_title, end_column=3)
    ws.cell(row=row_max_bm_title, column=2, value="Maximum change in Bone Marrow  Blasts")

    row_max_bm_fold = row_max_bm_title + 1
    ws.merge_cells(start_row=row_max_bm_fold, start_column=2, end_row=row_max_bm_fold, end_column=3)
    ws.cell(row=row_max_bm_fold, column=2, value="fold change")
    for col_idx, grp in zip(cols, cohort_names + ["__OVERALL__"]):
        keys = ok_keys if grp == "__OVERALL__" else set(cohorts[grp])
        ws.cell(row=row_max_bm_fold, column=col_idx, value=max_fold_reduction_bm(keys)).alignment = Alignment(horizontal="center")

    # Section 2: CD123+ BMB (BM*CD123)
    row_cd_title = row_max_bm_fold + 1
    ws.merge_cells(start_row=row_cd_title, start_column=2, end_row=row_cd_title, end_column=3)
    ws.cell(row=row_cd_title, column=2, value="Frequency of CD123+ BMB (%, range)")

    for i, v in enumerate(visit_list, start=row_cd_title+1):
        ws.cell(row=i, column=3, value=str(v))
        for col_idx, grp in zip(cols, cohort_names + ["__OVERALL__"]):
            keys = ok_keys if grp == "__OVERALL__" else set(cohorts[grp])
            m, mn, mx = mean_min_max_product(keys, v)
            text = "-" if m is None else f"{m}%\n({mn}–{mx})"
            ws.cell(row=i, column=col_idx, value=text).alignment = Alignment(wrap_text=True, vertical="top")

    # Max fold change in CD123+ BMB
    row_max_cd_title = row_cd_title + 1 + n_vis
    ws.merge_cells(start_row=row_max_cd_title, start_column=2, end_row=row_max_cd_title, end_column=3)
    ws.cell(row=row_max_cd_title, column=2, value="Maximum fold change in CD123+ Bone Marrow Blasts")

    row_max_cd_fold = row_max_cd_title + 1
    ws.merge_cells(start_row=row_max_cd_fold, start_column=2, end_row=row_max_cd_fold, end_column=3)
    ws.cell(row=row_max_cd_fold, column=2, value="fold change")
    for col_idx, grp in zip(cols, cohort_names + ["__OVERALL__"]):
        keys = ok_keys if grp == "__OVERALL__" else set(cohorts[grp])
        ws.cell(row=row_max_cd_fold, column=col_idx, value=max_fold_reduction_product(keys)).alignment = Alignment(horizontal="center")

    # Merge Column A from A3 down to the last C row of the CD123 section
    last_c_row = row_cd_title + max(n_vis, 1)
    ws.merge_cells(start_row=3, start_column=1, end_row=last_c_row, end_column=1)
    ws["A3"].value = "Bone Marrow\nBlasts"
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # ----------------- TP53 section (total + categories) -----------------
    tp53_start = row_max_cd_fold + 2

    # Merge 12 cells in Column A with "TP53"
    a1 = tp53_start
    a2 = tp53_start + 11
    ws.merge_cells(start_row=a1, start_column=1, end_row=a2, end_column=1)
    ws.cell(row=a1, column=1, value="TP53").alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    # ---- Total TP53 Mutated Pts (values on second row) ----
    ws.cell(row=tp53_start,   column=2, value="Total TP53 Mutated Pts")
    ws.cell(row=tp53_start+1, column=3, value="Overall count, average VAF (%)")

    col_tp53 = find_col(input_df, ["TP53FINALRESULT","TP53 FINAL RESULT","TP53FINALRESULT__"])  # tolerate weird headers
    def _mutated_count_text(keys: set) -> str:
        if not col_tp53:
            return "0 (-)"
        sub_all = input_df[input_df[col_subject].map(_keyize_id).isin(keys)].copy()
        sub_all = sub_all[~sub_all[col_visit].map(is_excluded_visit)]

        # keep only subjects with ≥1 valid C#D# non-baseline visit
        keep_keys = set()
        for k, g in sub_all.groupby(sub_all[col_subject].map(_keyize_id)):
            if has_valid_timepoint_nonbaseline(g[col_visit]):
                keep_keys.add(k)
        if not keep_keys:
            return "0 (-)"

        sub = input_df[input_df[col_subject].map(_keyize_id).isin(keep_keys)].copy()
        lab = sub[col_tp53].astype(str).str.replace("\u00A0"," ").str.strip().str.lower()
        # be slightly permissive on labels
        mutated_mask = lab.str.fullmatch(r"(positive|pos|detected)")
        mutated_keys = set(sub.loc[mutated_mask, col_subject].map(_keyize_id).dropna().unique().tolist())
        if not mutated_keys:
            return "0 (-)"

        # average across all TP53VAF* values from mutated subjects
        _, vaf_cols, _ = _detect_mut_and_vaf_cols(input_df)
        vafs = []
        if vaf_cols:
            sub2 = input_df[input_df[col_subject].map(_keyize_id).isin(mutated_keys)].copy()
            for vc in vaf_cols:
                vals = pd.to_numeric(sub2[vc].map(_to_percent), errors="coerce").dropna().tolist()
                vafs.extend(vals)
        if not vafs:
            return f"{len(mutated_keys)} (-)"
        mean_v = round(sum(vafs)/len(vafs), 1)
        return f"{len(mutated_keys)} ({mean_v}%)"

    for col_idx, grp in zip(cols, cohort_names + ["__OVERALL__"]):
        keys = ok_keys if grp == "__OVERALL__" else set(cohorts.get(grp, []))
        ws.cell(row=tp53_start+1, column=col_idx, value=_mutated_count_text(keys)).alignment = Alignment(horizontal="center")

    # ---- Precompute per-cohort mutation dataframes for category blocks ----
    def _mut_df(keys: set) -> pd.DataFrame:
        return _cohort_mutations_df(
            input_df=input_df,
            keys=keys,
            col_subject=col_subject,
            col_visit=col_visit,
            dict_path=tp53_dict_path
        )

    per_cohort_mut = {}
    for grp in cohort_names:
        per_cohort_mut[grp] = _mut_df(set(cohorts[grp]))
    per_cohort_mut["__OVERALL__"] = _mut_df(ok_keys)

    # ---- Generic writer for a 2-row category block ----
    def _append_category_block(start_row: int, label_b: str, label_c: str, mask_builder) -> int:
        ws.cell(row=start_row,   column=2, value=label_b)
        ws.cell(row=start_row+1, column=3, value=label_c)
        for col_idx, grp in zip(cols, cohort_names + ["__OVERALL__"]):
            dfm = per_cohort_mut.get(grp)
            if dfm is None or dfm.empty:
                text = "0 (-)"
            else:
                mask = mask_builder(dfm)
                # If mask length mismatches (e.g., empty), guard it
                if mask is None or len(mask) != len(dfm):
                    text = "0 (-)"
                else:
                    text = _count_mean_text(dfm, mask)
            ws.cell(row=start_row+1, column=col_idx, value=text).alignment = Alignment(horizontal="center")
        return start_row + 2

    # ---- Missense block ----
    row_ptr = tp53_start + 2
    row_ptr = _append_category_block(
        start_row=row_ptr,
        label_b="Mutation Type (Missense)",
        label_c="Count, average VAF (%)",
        mask_builder=lambda df: df["Effect"].astype(str).str.lower().eq("missense")
    )

    # ---- Truncating block (FS, nonsense, splice) ----
    trunc_set = {"fs","nonsense","splice"}
    row_ptr = _append_category_block(
        start_row=row_ptr,
        label_b="Mutation Type (Truncating)",
        label_c="Count, average VAF (%)",
        mask_builder=lambda df: df["Effect"].astype(str).str.lower().isin(trunc_set)
    )

    # ---- Hotspot block ----
    hotspot_label_c = "Arg175, Gly245, Arg248, Arg249, Arg273, or Arg282\nCount, average VAF (%)"
    row_ptr = _append_category_block(
        start_row=row_ptr,
        label_b="Hotspot Mutations",
        label_c=hotspot_label_c,
        mask_builder=lambda df: df["Hotspot"].astype(str).str.lower().eq("yes")
    )

    # ---- High VAF ≥40% ----
    row_ptr = _append_category_block(
        start_row=row_ptr,
        label_b="High VAF Clones ≥40%",
        label_c="Count, average VAF (%)",
        mask_builder=lambda df: pd.to_numeric(df["vaf"], errors="coerce").ge(40.0)
    )

    # ---- Low VAF <40% ----
    row_ptr = _append_category_block(
        start_row=row_ptr,
        label_b="Low VAF Clones <40%",
        label_c="Count, average VAF (%)",
        mask_builder=lambda df: pd.to_numeric(df["vaf"], errors="coerce").lt(40.0)
    )

    # widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 36
    for c in cols:
        ws.column_dimensions[get_column_letter(c)].width = 22

    # Save
    out_dir = os.path.dirname(out_path)
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)

# ----------------- GUI -----------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        try:
            os.chdir(os.path.expanduser("~"))
        except Exception:
            pass

        self.title("FDA combo table — Builder")
        self.geometry("980x520")
        self.resizable(False, False)

        self.last_dir = os.path.expanduser("~")
        self.input_path = tk.StringVar()
        self.dose_path  = tk.StringVar()
        self.tp53_path  = tk.StringVar()
        self.out_dir    = tk.StringVar()
        self.out_name   = tk.StringVar(value="FDA_combo_output.xlsx")

        pad = {"padx": 12, "pady": 8}
        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, **pad)
        frm.columnconfigure(1, weight=1)

        ttk.Label(frm, text="Input data file (with SUBJECT, VISIT, BMBLASTS/CD123)").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.input_path, width=80).grid(row=0, column=1, sticky="we", padx=(8,8))
        ttk.Button(frm, text="Browse…", command=self.browse_input).grid(row=0, column=2, sticky="e")

        ttk.Label(frm, text="Patient dose dictionary (Patient Number ↔ Treatment Group)").grid(row=1, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.dose_path, width=80).grid(row=1, column=1, sticky="we", padx=(8,8))
        ttk.Button(frm, text="Browse…", command=self.browse_dose).grid(row=1, column=2, sticky="e")

        ttk.Label(frm, text="TP53 mutation dictionary (CSV/XLSX, index_col=0)").grid(row=2, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.tp53_path, width=80).grid(row=2, column=1, sticky="we", padx=(8,8))
        ttk.Button(frm, text="Browse…", command=self.browse_tp53).grid(row=2, column=2, sticky="e")

        ttk.Label(frm, text="Output folder").grid(row=3, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.out_dir, width=80).grid(row=3, column=1, sticky="we", padx=(8,8))
        ttk.Button(frm, text="Choose…", command=self.browse_outdir).grid(row=3, column=2, sticky="e")

        ttk.Label(frm, text="Output file name (.xlsx)").grid(row=4, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.out_name, width=40).grid(row=4, column=1, sticky="w", padx=(8,8))

        btns = ttk.Frame(frm)
        btns.grid(row=6, column=0, columnspan=3, sticky="e", pady=(16,0))
        ttk.Button(btns, text="Quit", command=self.destroy).pack(side="right", padx=(8,0))
        ttk.Button(btns, text="Run", command=self.run).pack(side="right")

    def browse_input(self):
        path = filedialog.askopenfilename(title="Select Input Data", initialdir=self.last_dir, filetypes=ALLOWED_ANY_DATA)
        if path:
            self.last_dir = os.path.dirname(path)
            self.input_path.set(path)

    def browse_dose(self):
        path = filedialog.askopenfilename(title="Select Patient Dose Dictionary", initialdir=self.last_dir, filetypes=ALLOWED_ANY_DATA)
        if path:
            self.last_dir = os.path.dirname(path)
            self.dose_path.set(path)

    def browse_tp53(self):
        path = filedialog.askopenfilename(title="Select TP53 Mutation Dictionary (optional)", initialdir=self.last_dir, filetypes=ALLOWED_ANY_DATA)
        if path:
            self.last_dir = os.path.dirname(path)
            self.tp53_path.set(path)

    def browse_outdir(self):
        path = filedialog.askdirectory(title="Select Output Folder", initialdir=self.last_dir)
        if path:
            self.last_dir = path
            self.out_dir.set(path)

    def run(self):
        in_path   = self.input_path.get().strip()
        dose_path = self.dose_path.get().strip()
        tp53_path = self.tp53_path.get().strip() or None
        out_dir   = self.out_dir.get().strip()
        out_name  = self.out_name.get().strip()

        if not in_path or not os.path.isfile(in_path):
            messagebox.showerror("Missing input", "Please select a valid INPUT data file."); return
        if not dose_path or not os.path.isfile(dose_path):
            messagebox.showerror("Missing dose dict", "Please select a valid patient dose dictionary file."); return
        if not out_dir or not os.path.isdir(out_dir):
            messagebox.showerror("Missing output folder", "Please choose a valid output folder."); return
        if out_name == "":
            messagebox.showerror("Missing output name", "Please provide an output file name (e.g., FDA_combo_output.xlsx)."); return
        if not out_name.lower().endswith(".xlsx"):
            out_name += ".xlsx"
        out_path = os.path.join(out_dir, out_name)

        try:
            generate_fda_combo_excel(in_path, dose_path, tp53_path, out_path)
        except Exception as e:
            messagebox.showerror("Processing error", str(e)); return

        messagebox.showinfo("Done", f"Excel written to:\n{out_path}")

# ---------- entry ----------
if __name__ == "__main__":
    App().mainloop()
